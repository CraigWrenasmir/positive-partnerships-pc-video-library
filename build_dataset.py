from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path("/Users/craigsmith/Desktop/Workshop Facilitator Guides")
WORKBOOK_PATH = BASE_DIR / "PC_video_inventory.xlsx"
TRANSCRIPTS_DIR = BASE_DIR / "2 Day PC Transcripts"
OUTPUT_PATH = BASE_DIR / "pc-video-library" / "data.json"


def slugify(value: str) -> str:
    value = value.lower().replace("&", "and")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def parse_transcript(text: str) -> list[dict[str, str]]:
    blocks: list[dict[str, str]] = []
    current_speaker = ""
    current_lines: list[str] = []

    def flush() -> None:
        nonlocal current_speaker, current_lines
        content = " ".join(line.strip() for line in current_lines if line.strip()).strip()
        if content:
            blocks.append(
                {
                    "speaker": current_speaker or "Narration",
                    "text": content,
                }
            )
        current_speaker = ""
        current_lines = []

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            flush()
            continue
        speaker_match = re.match(r"^([A-Za-z0-9 .'()-]+):\s*$", line)
        if speaker_match:
            flush()
            current_speaker = speaker_match.group(1)
            continue
        current_lines.append(line)

    flush()
    return blocks


def transcript_payload(filename: str) -> dict | None:
    if not filename:
        return None
    transcript_path = TRANSCRIPTS_DIR / filename
    if not transcript_path.exists():
        return None
    raw_text = transcript_path.read_text(encoding="utf-8").strip()
    return {
        "file": filename,
        "plainText": raw_text,
        "blocks": parse_transcript(raw_text),
    }


def main() -> None:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    unique_sheet = workbook["Unique Videos"]
    mapping_sheet = workbook["Workshop Mapping"]

    videos: list[dict] = []
    videos_by_id: dict[str, dict] = {}

    for row in unique_sheet.iter_rows(min_row=7, values_only=True):
        (
            video_id,
            catalog_title,
            speaker_or_source,
            duration,
            slide_titles,
            transcript_status,
            transcript_file,
            used_in_1_day_pc,
            used_in_2_day_pc,
            notes,
            match_note,
        ) = row
        if not video_id:
            continue
        video = {
            "id": video_id,
            "slug": slugify(f"{video_id}-{catalog_title}"),
            "title": catalog_title,
            "speakerOrSource": speaker_or_source or "",
            "duration": duration or "",
            "slideTitles": [item.strip() for item in str(slide_titles or "").split("|") if item.strip()],
            "transcriptStatus": transcript_status or "missing",
            "transcriptFile": transcript_file or "",
            "notes": notes or "",
            "matchNote": match_note or "",
            "summaryStatus": "pending",
            "summary": None,
            "transcript": transcript_payload(transcript_file or ""),
        }
        videos.append(video)
        videos_by_id[video_id] = video

    workshop_map = {
        "1-day-pc": {"id": "1-day-pc", "name": "1 Day PC", "videos": []},
        "2-day-pc": {"id": "2-day-pc", "name": "2 Day PC", "videos": []},
    }

    for row in mapping_sheet.iter_rows(min_row=2, values_only=True):
        (
            workshop,
            video_id,
            catalog_title,
            speaker_or_source,
            duration,
            slide_title,
            pdf_file,
            pdf_page,
            transcript_status,
            transcript_file,
            notes,
            match_note,
        ) = row
        if workshop not in {"1 Day PC", "2 Day PC"} or not video_id:
            continue
        workshop_id = "1-day-pc" if workshop == "1 Day PC" else "2-day-pc"
        workshop_map[workshop_id]["videos"].append(
            {
                "videoId": video_id,
                "title": catalog_title,
                "slideTitle": slide_title or "",
                "pdfFile": pdf_file or "",
                "pdfPage": int(pdf_page) if pdf_page else None,
                "transcriptStatus": transcript_status or "missing",
                "transcriptFile": transcript_file or "",
                "notes": notes or "",
                "matchNote": match_note or "",
            }
        )

    dataset = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "workshops": list(workshop_map.values()),
        "videos": videos,
    }

    OUTPUT_PATH.write_text(json.dumps(dataset, indent=2, ensure_ascii=True), encoding="utf-8")


if __name__ == "__main__":
    main()
